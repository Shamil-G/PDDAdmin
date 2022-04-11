import app_config as cfg
from db_oracle.connect import get_connection
from openpyxl import load_workbook
import datetime
import os.path
from werkzeug.security import generate_password_hash
from flask import g
import cx_Oracle


def load_operators(file_name):
    s_now = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")

    if cfg.os == 'unix':
       file_path = cfg.UPLOAD_PATH + '/' + file_name
    else:
       file_path = cfg.UPLOAD_PATH + '\\' + file_name

    # Нормируем путь к файлу по слэшам
    path = os.path.normpath(file_path)

    print(f"Загрузка стартовала: {s_now}, file_name: {file_path}")

    if not os.path.isfile(file_path):
        print("File not exists: " + str(os.path.isfile(file_path)))
        return file_name
    print("Load Theme with Excel file: " + str(os.path.isfile(file_path)))

    wb = load_workbook(path)
    sheet_number = len(wb.worksheets)
    print("Книга загружена: " + " : " + path)
    sheet = wb.active

    print("Подключаем БД")

    con = get_connection()
    cursor = con.cursor()
    # Создадим новое задание
    id_quest = 0
    id_prev_quest = -1
    order_num = 0
    for sheet in wb.worksheets:
        for i in range(2, sheet.max_row+1):
            id_rec = sheet.cell(row=i, column=1).value
            descr = sheet.cell(row=i, column=2).value
            fio = sheet.cell(row=i, column=3).value
            iin = str(sheet.cell(row=i, column=4).value).rjust(12, '0')
            phone = sheet.cell(row=i, column=5).value
            username = sheet.cell(row=i, column=6).value
            password = sheet.cell(row=i, column=7).value

            if iin and username and password:
                if not id_rec:
                    break
                fio_split = fio.split(' ')
                fname = fio_split[1]
                flast_name = fio_split[0]
                if len(fio_split) == 3:
                    fmiddle_name = fio_split[2]
                else:
                    fmiddle_name = ''

                hash_pwd = generate_password_hash(password)
                message = cursor.var(cx_Oracle.DB_TYPE_NVARCHAR)
                # cmd = "insert into cop.users (id_user, active, phone, iin, date_op, username, password, " \
                #       "lastname, name, middlename, descr) " \
                #       f"values (seq_persons.nextval, 'Y', '{phone}', '{iin}', '{s_now}', '{username}', '{password}', " \
                #       f"'{flast_name}', '{fname}', '{fmiddle_name}', '{descr}')"
                print(f"iin: {iin}, phone: {phone}, fname: {fname}, flast_name: {flast_name}, fmiddle_name: {fmiddle_name}")
                cursor.callproc('cop.cop.new_user2', [username, hash_pwd, int(g.user.id_user), iin, phone,
                                fname, flast_name, fmiddle_name, f'Ustudy: {descr}', message])
                # cursor.execute(cmd)

    con.commit()
    con.close()
    now = datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    print(f"Загрузка завершена: {now}")
    return


if __name__ == "__main__":
    print("Тестируем загрузку operators from Excel!")
    load_operators('operators.xlsx')
