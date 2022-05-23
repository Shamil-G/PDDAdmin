from db_oracle.connect import get_connection
import main_app as cfg
from openpyxl import load_workbook
import datetime
import os.path


def load_partition_questions(id_task, lang, partition_number, subpartition_number, file_name):
    theme_number = 2
    s_now = datetime.datetime.now()
    if cfg.os == 'unix':
       file_path = cfg.UPLOAD_PATH + '/' + lang + '/' + file_name
    else:
       file_path = cfg.UPLOAD_PATH + '\\' + lang + '\\' + file_name

    # Нормируем путь к файлу по слэшам
    path = os.path.normpath(file_path)

    print("Загрузка стартовала: " + s_now.strftime("%d-%m-%Y %H:%M:%S") + ' : ' + file_name + ' : ' + file_path)

    if not os.path.isfile(file_path):
        print("File not exists: " + str(os.path.isfile(file_path)))
        return file_name
    print("Load Theme with Excel file: " + str(os.path.isfile(file_path)))

    wb = load_workbook(path)
    print("Книга загружена: " + path)
    sheet = wb.active

    print("Подключаем БД")

    con = get_connection()
    cursor = con.cursor()
    # Создадим новое задание
    # file_split = os.path.splitext(file_name)
    # id_theme = cursor.callfunc('admin.theme_new', int, (id_task, file_split[0]))
    # if not id_theme:
    #     print('Ошибка регистрации нового задания...')
    id_quest = 0
    id_prev_quest = -1
    order_num = 0
    for i in range(2, sheet.max_row+1):
        id_curr_quest = sheet.cell(row=i, column=1).value
        quest = sheet.cell(row=i, column=2).value
        correctly = sheet.cell(row=i, column=3).value
        answer = sheet.cell(row=i, column=4).value
        url_image = sheet.cell(row=i, column=5).value
        order_num = order_num + 1
        if not quest:
            break
        if id_curr_quest != id_prev_quest:
            id_quest = id_quest + 1
            order_num = 1
            id_question = cursor.callfunc("pdd_testing.load_questions.add_question", str,
                                          [id_task, theme_number,
                                           partition_number, subpartition_number, id_quest, url_image, quest])

        cursor.callproc("pdd_testing.load_questions.add_answer", [id_question, order_num, correctly, answer])
        id_prev_quest = id_curr_quest

    con.commit()
    con.close()
    now = datetime.datetime.now()
    print("Загрузка завершена: " + now.strftime("%d-%m-%Y %H:%M:%S"))
    return


def load_task(id_task, lang):
    part = 1
    for sub_part in range(1, 15):
        f_name = str(sub_part) + ' подраздел.xlsx'
        load_partition_questions(id_task, lang, part, sub_part, f_name)

    f_name = 'ОБД.xlsx'
    part = 2
    load_partition_questions(id_task, lang, part, 0, f_name)

    f_name = 'Медицина.xlsx'
    part = 3
    load_partition_questions(id_task, lang, part, 0, f_name)

    f_name = 'Административная ответственность.xlsx'
    part = 4
    load_partition_questions(id_task, lang, part, 0, f_name)

    f_name = 'ПДДАА1В1.xlsx'
    part = 5
    sub_part = 1
    load_partition_questions(id_task, lang, part, sub_part, f_name)

    f_name = 'ПДДВВЕ.xlsx'
    part = 5
    sub_part = 2
    load_partition_questions(id_task, lang, part, sub_part, f_name)

    f_name = 'ПДДС1С.xlsx'
    part = 5
    sub_part = 3
    load_partition_questions(id_task, lang, part, sub_part, f_name)

    f_name = 'ПДДD1DТb.xlsx'
    part = 5
    sub_part = 4
    load_partition_questions(id_task, lang, part, sub_part, f_name)

    f_name = 'ПДДC1ECED1EDE.xlsx'
    part = 5
    sub_part = 5
    load_partition_questions(id_task, lang, part, sub_part, f_name)

    f_name = 'ПДД Tm.xlsx'
    part = 5
    sub_part = 6
    load_partition_questions(id_task, lang, part, sub_part, f_name)

    f_name = 'СПОБДА1АВ1.xlsx'
    part = 6
    sub_part = 1
    load_partition_questions(id_task, lang, part, sub_part, f_name)

    f_name = 'СПОБДС1С.xlsx'
    part = 6
    sub_part = 3
    load_partition_questions(id_task, lang, part, sub_part, f_name)

    f_name = 'СПОБДD1DTb.xlsx'
    part = 6
    sub_part = 4
    load_partition_questions(id_task, lang, part, sub_part, f_name)

    f_name = 'СПОБДC1ECED1EDE.xlsx'
    part = 6
    sub_part = 5
    load_partition_questions(id_task, lang, part, sub_part, f_name)

    f_name = 'СПОБДTm.xlsx'
    part = 6
    sub_part = 6
    load_partition_questions(id_task, lang, part, sub_part, f_name)


if __name__ == "__main__":
    load_task(1, 'ru')
    print("--------------> Загружены вопросы на русском языке!")
    load_task(2, 'kz')
    print("--------------> Загружены вопросы на казахском языке!")
    load_task(3, 'en')
    print("--------------> Загружены вопросы на английском языке!")
