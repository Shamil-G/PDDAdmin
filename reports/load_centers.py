from db_oracle.connect import get_connection
import config as cfg
from openpyxl import load_workbook
import datetime
import os.path


def load_ip(file_name):
    s_now = datetime.datetime.now()
    if cfg.os == 'unix':
       file_path = cfg.UPLOAD_PATH + '/' + file_name
    else:
       file_path = cfg.UPLOAD_PATH + '\\' + file_name

    # Нормируем путь к файлу по слэшам
    path = os.path.normpath(file_path)

    print("Загрузка стартовала: " + s_now.strftime("%d-%m-%Y %H:%M:%S") + ' : ' + file_name + ' : ' + file_path)

    if not os.path.isfile(file_path):
        print("File not exists: " + str(os.path.isfile(file_path)))
        return file_name
    print("Load Theme with Excel file: " + str(os.path.isfile(file_path)))

    wb = load_workbook(path)
    sheet_number = len(wb.worksheets)
    print("Книга загружена: " + " : " + path)
    sheet = wb.active

    print("Подключаем БД")

    con = get_connection()
    cursor = con.cursor()
    # Создадим новое задание
    id_quest = 0
    id_prev_quest = -1
    order_num = 0
    for sheet in wb.worksheets:
        for i in range(2, sheet.max_row+1):
            id = sheet.cell(row=i, column=1).value
            ip = sheet.cell(row=i, column=2).value
            mac = sheet.cell(row=i, column=3).value
            name = sheet.cell(row=i, column=4).value
            code = sheet.cell(row=i, column=5).value
            order_num = order_num + 1
            if not ip:
                break
            cmd = "insert into list_ip (id, ip_addr, mac, name, code) " \
                  "values ( " + str(id) + ", '" + ip + "', '" + mac + "', '" + name + "', '" + code + "')"
            print('+++ CMD: ' + cmd)
            cursor.execute(cmd)

    con.commit()
    con.close()
    now = datetime.datetime.now()
    print("Загрузка завершена: " + now.strftime("%d-%m-%Y %H:%M:%S"))
    return


def load_center(file_name):
    s_now = datetime.datetime.now()
    if cfg.os == 'unix':
       file_path = cfg.UPLOAD_PATH + '/' + file_name
    else:
       file_path = cfg.UPLOAD_PATH + '\\' + file_name

    # Нормируем путь к файлу по слэшам
    path = os.path.normpath(file_path)

    print("Загрузка стартовала: " + s_now.strftime("%d-%m-%Y %H:%M:%S") + ' : ' + file_name + ' : ' + file_path)

    if not os.path.isfile(file_path):
        print("File not exists: " + str(os.path.isfile(file_path)))
        return file_name
    print("Load Theme with Excel file: " + str(os.path.isfile(file_path)))

    wb = load_workbook(path)
    print("Книга загружена: " + " : " + path)
    sheet = wb.active

    print("Подключаем БД")

    con = get_connection()
    cursor = con.cursor()
    # Создадим новое задание
    id_quest = 0
    id_prev_quest = -1
    order_num = 0
    for sheet in wb.worksheets:
        for i in range(2, sheet.max_row+1):
            id = sheet.cell(row=i, column=1).value
            region = sheet.cell(row=i, column=2).value
            code = sheet.cell(row=i, column=3).value
            short_rus = sheet.cell(row=i, column=4).value
            full_rus = sheet.cell(row=i, column=5).value
            short_kz = sheet.cell(row=i, column=6).value
            full_kz = sheet.cell(row=i, column=7).value
            name_exec_rus = sheet.cell(row=i, column=8).value
            name_exec_kaz = sheet.cell(row=i, column=9).value
            exec_code = sheet.cell(row=i, column=10).value
            order_num = order_num + 1
            if not id:
                break
            cmd = "insert into svod (id, region, code, short_rus, full_rus, short_kz, full_kz, " \
                "name_exec_rus, name_exec_kaz, exec_code) " \
                "values ( " + str(id) + ", '" + region + "', '" + code + "', '" + short_rus + "', '" \
                + full_rus + "', '" + short_kz + "', '" + full_kz + "', '" + name_exec_rus + "', '" + \
                name_exec_kaz + "', '" + exec_code + "')"
            print('+++ CMD: ' + str(i) + ' : ' + cmd)
            cursor.execute(cmd)

    con.commit()
    con.close()
    now = datetime.datetime.now()
    print("Загрузка завершена: " + now.strftime("%d-%m-%Y %H:%M:%S"))
    return


if __name__ == "__main__":
    print("Тестируем загрузку Excel!")
    # load_ip('ip.xlsx')
    load_center('svod.xlsx')
